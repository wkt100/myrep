"""Parse user input from text, PDF, CSV, or Excel files."""

import csv
import os
from pathlib import Path
from typing import Optional


def read_text(path: str) -> str:
    """Read plain text file."""
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def read_pdf(path: str) -> str:
    """Extract text from PDF using PyMuPDF."""
    import fitz

    text_parts = []
    with fitz.open(path) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def read_csv(path: str) -> str:
    """Read CSV file and format as markdown table."""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)

    if not rows:
        return ""

    # Build markdown table
    lines = []
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def read_xlsx(path: str) -> str:
    """Read Excel file and format as markdown table."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        text_parts.append(f"## Sheet: {sheet_name}")
        lines = []
        lines.append("| " + " | ".join(str(v or "") for v in rows[0]) + " |")
        lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
        for row in rows[1:]:
            lines.append("| " + " | ".join(str(v or "") for v in row) + " |")
        text_parts.extend(lines)
    wb.close()
    return "\n".join(text_parts)


SUPPORTED_TYPES = {
    ".txt": "text",
    ".md": "text",
    ".pdf": "pdf",
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".xls": "xlsx",
}


def detect_type(path: str) -> str:
    """Detect input type from file extension."""
    ext = Path(path).suffix.lower()
    if ext in SUPPORTED_TYPES:
        return SUPPORTED_TYPES[ext]
    raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(SUPPORTED_TYPES)}")


def parse_input(source: str, source_type: Optional[str] = None) -> str:
    """Parse input from a file path or raw text.

    Args:
        source: File path or raw text string.
        source_type: One of 'text', 'pdf', 'csv', 'xlsx'.
                     If None, auto-detect from file extension (or treat as text).

    Returns:
        Extracted text content.
    """
    # If source_type is given, treat source as a file path of that type
    if source_type:
        if source_type == "text":
            return read_text(source)
        elif source_type == "pdf":
            return read_pdf(source)
        elif source_type == "csv":
            return read_csv(source)
        elif source_type == "xlsx":
            return read_xlsx(source)
        else:
            raise ValueError(f"Unknown source_type: {source_type}")

    # Auto-detect: if file exists, detect by extension
    if os.path.isfile(source):
        stype = detect_type(source)
        if stype == "text":
            return read_text(source)
        elif stype == "pdf":
            return read_pdf(source)
        elif stype == "csv":
            return read_csv(source)
        elif stype == "xlsx":
            return read_xlsx(source)

    # Treat as raw text
    return source


def summarize_content(text: str, max_len: int = 500) -> str:
    """Show first portion of content for user preview."""
    if len(text) <= max_len:
        return text
    return text[:max_len] + f"\n\n... (共 {len(text)} 字符，仅显示前 {max_len} 字符)"
